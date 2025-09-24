import logging
import pandas as pd
import io
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from .models import RawFileStructure, RawAnalysisResult, FileType
from .gpt5_column_analyzer import GPT5ColumnAnalyzer

logger = logging.getLogger(__name__)

class RawFileAnalyzer:
    """GPT-5 powered raw file analysis for intelligent pre-processing"""
    
    def __init__(self):
        """Initialize raw file analyzer with GPT-5"""
        try:
            self.gpt5_analyzer = GPT5ColumnAnalyzer()
            logger.info("Raw File Analyzer initialized with GPT-5")
        except Exception as e:
            logger.warning(f"GPT-5 initialization failed: {str(e)}")
            self.gpt5_analyzer = None
    
    async def analyze_raw_file_structure(
        self, 
        file_content: bytes, 
        file_type: FileType, 
        filename: str
    ) -> RawAnalysisResult:
        """
        Analyze raw file structure using GPT-5 before Docling/pandas processing
        
        Returns intelligent recommendations for:
        - Sheet selection (Excel files)
        - Header row detection
        - Data start position
        - Column type hints
        - German accounting patterns
        """
        try:
            if file_type == FileType.XLSX:
                return await self._analyze_excel_structure(file_content, filename)
            elif file_type == FileType.CSV:
                return await self._analyze_csv_structure(file_content, filename)
            else:
                # For PDF files, return basic structure
                return RawAnalysisResult(
                    file_structure=RawFileStructure(confidence=0.5),
                    content_preview=[],
                    processing_hints={"requires_ocr": True},
                    analysis_confidence=0.5
                )
                
        except Exception as e:
            logger.error(f"Raw file analysis failed: {str(e)}")
            return RawAnalysisResult(
                file_structure=RawFileStructure(confidence=0.0),
                content_preview=[],
                processing_hints={"error": str(e)},
                analysis_confidence=0.0
            )
    
    async def _analyze_excel_structure(self, file_content: bytes, filename: str) -> RawAnalysisResult:
        """Analyze Excel file structure using GPT-5"""
        try:
            # Load workbook to get sheet names and sample content
            workbook = load_workbook(io.BytesIO(file_content), read_only=True)
            sheet_names = workbook.sheetnames
            
            # Get preview of each sheet (first 20 rows)
            sheet_previews = {}
            for sheet_name in sheet_names[:5]:  # Limit to first 5 sheets
                try:
                    df_preview = pd.read_excel(
                        io.BytesIO(file_content),
                        sheet_name=sheet_name,
                        header=None,
                        nrows=20
                    )
                    # Convert to string representation for GPT-5 analysis
                    preview_text = self._dataframe_to_preview_text(df_preview, sheet_name)
                    sheet_previews[sheet_name] = preview_text
                except Exception as e:
                    logger.warning(f"Could not preview sheet {sheet_name}: {str(e)}")
                    continue
            
            workbook.close()
            
            # Use GPT-5 to analyze sheet structure
            if self.gpt5_analyzer:
                analysis = await self.gpt5_analyzer.analyze_raw_excel_structure(
                    sheet_names, sheet_previews, filename
                )
                return analysis
            else:
                # Fallback analysis without GPT-5
                return await self._fallback_excel_analysis(sheet_names, sheet_previews)
                
        except Exception as e:
            logger.error(f"Excel structure analysis failed: {str(e)}")
            raise
    
    async def _analyze_csv_structure(self, file_content: bytes, filename: str) -> RawAnalysisResult:
        """Analyze CSV file structure using GPT-5"""
        try:
            # Get first 50 lines for analysis
            text_content = file_content.decode('utf-8', errors='ignore')
            lines = text_content.split('\n')[:50]
            
            # Try different CSV delimiters
            delimiter_candidates = [',', ';', '\t', '|']
            best_delimiter = ','
            max_columns = 0
            
            for delim in delimiter_candidates:
                try:
                    df_test = pd.read_csv(io.StringIO('\n'.join(lines[:10])), 
                                        sep=delim, header=None)
                    if len(df_test.columns) > max_columns:
                        max_columns = len(df_test.columns)
                        best_delimiter = delim
                except:
                    continue
            
            # Create preview with best delimiter
            preview_df = pd.read_csv(
                io.StringIO('\n'.join(lines)),
                sep=best_delimiter,
                header=None,
                nrows=30
            )
            
            preview_text = self._dataframe_to_preview_text(preview_df, "CSV Data")
            
            # Use GPT-5 to analyze CSV structure
            if self.gpt5_analyzer:
                analysis = await self.gpt5_analyzer.analyze_raw_csv_structure(
                    lines, preview_text, filename, best_delimiter
                )
                return analysis
            else:
                # Fallback analysis without GPT-5
                return await self._fallback_csv_analysis(lines, best_delimiter)
                
        except Exception as e:
            logger.error(f"CSV structure analysis failed: {str(e)}")
            raise
    
    def _dataframe_to_preview_text(self, df: pd.DataFrame, sheet_name: str) -> str:
        """Convert DataFrame to text preview for GPT-5 analysis"""
        preview_lines = [f"=== {sheet_name} ==="]
        
        for idx, row in df.iterrows():
            if idx >= 20:  # Limit preview length
                break
            row_values = []
            for val in row:
                if pd.isna(val):
                    row_values.append("[EMPTY]")
                else:
                    row_values.append(str(val)[:50])  # Truncate long values
            preview_lines.append(f"Row {idx+1}: {' | '.join(row_values)}")
        
        return '\n'.join(preview_lines)
    
    async def _fallback_excel_analysis(self, sheet_names: List[str], sheet_previews: Dict[str, str]) -> RawAnalysisResult:
        """Fallback Excel analysis without GPT-5"""
        # Simple heuristics for sheet selection
        preferred_sheet = sheet_names[0]  # Default to first sheet
        
        # Look for sheets with common accounting names
        accounting_keywords = ['tb', 'trial', 'balance', 'saldo', 'bwa', 'bilanz']
        for sheet_name in sheet_names:
            if any(keyword in sheet_name.lower() for keyword in accounting_keywords):
                preferred_sheet = sheet_name
                break
        
        return RawAnalysisResult(
            file_structure=RawFileStructure(
                recommended_sheet=preferred_sheet,
                header_row=1,
                data_start_row=2,
                column_hints={},
                detected_patterns=['fallback_analysis'],
                confidence=0.6,
                recommendations=["Using fallback analysis - GPT-5 not available"]
            ),
            content_preview=list(sheet_previews.values())[:3],
            processing_hints={
                "fallback_mode": True,
                "recommended_sheet": preferred_sheet
            },
            analysis_confidence=0.6
        )
    
    async def _fallback_csv_analysis(self, lines: List[str], delimiter: str) -> RawAnalysisResult:
        """Fallback CSV analysis without GPT-5"""
        # Simple heuristics for header detection
        header_row = 0
        data_start_row = 1
        
        # Look for German accounting keywords in first few rows
        german_keywords = ['konto', 'saldo', 'betrag', 'bezeichnung']
        for i, line in enumerate(lines[:5]):
            if any(keyword in line.lower() for keyword in german_keywords):
                header_row = i
                data_start_row = i + 1
                break
        
        return RawAnalysisResult(
            file_structure=RawFileStructure(
                header_row=header_row,
                data_start_row=data_start_row,
                column_hints={},
                detected_patterns=['fallback_analysis', f'delimiter_{delimiter}'],
                confidence=0.6,
                recommendations=["Using fallback analysis - GPT-5 not available"]
            ),
            content_preview=lines[:10],
            processing_hints={
                "fallback_mode": True,
                "delimiter": delimiter
            },
            analysis_confidence=0.6
        )